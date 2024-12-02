import pandas as pd
from entities import *
import openpyxl as xl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import subprocess

import json

LEC_COLOR = '#a98be5'
LAB_COLOR = '#f7d483'

LESSON_TIME = [
    '9:00 - 10:20',
    '10:30 - 11:50',
    '12:00 - 13:20',
    '13:50 - 15:10',
    '15:20 - 16:40',
    '17:00 - 18:20',
    '18:30 - 19:50',
    '20:00 - 21:20'
]

WEEK_DAYS = [
    'ПОНЕДЕЛЬНИК',
    'ВТОРНИК',
    'СРЕДА',
    'ЧЕТВЕРГ',
    'ПЯТНИЦА',
    'СУББОТА',
]

GROUPS = [
    '1 РФ',
    '2 РФ',
    '3 РФ',
    '4 РФ',
    '8 РФ',
    '3 ФЭ',
    '2 АРИСТ',
    '8 АРИСТ',
    '4 КБ',
    '5 КБ',
    '6 КБ',
    '7 КБ',
    '1 ПИ',
    '5 ПИ',
    '6 ПИ',
    '7 ПИ'
]

f = open('data/data.json', 'r')
data = json.load(f)
f.close()
lessons = Lessons(**data)
schedule_df = pd.read_csv('data/example-1.csv')


def set_border(ws, cell_range, border, overwrite=False):
    def get_border_attr(b):
        return {
            'left': getattr(b, 'left'),
            'right': getattr(b, 'right'),
            'top': getattr(b, 'top'),
            'bottom': getattr(b, 'bottom'),
        }

    for row in ws[cell_range]:
        for cell in row:
            if overwrite:
                cell.border = border
            else:
                saved_border = get_border_attr(cell.border)
                new_border = get_border_attr(border)
                cell.border = Border(
                    left=new_border['left'] if new_border['left'] else saved_border['left'],
                    right=new_border['right'] if new_border['right'] else saved_border['right'],
                    top=new_border['top'] if new_border['top'] else saved_border['top'],
                    bottom=new_border['bottom'] if new_border['bottom'] else saved_border['bottom'],
                )


wb = xl.Workbook()
ws = wb.active
ws.freeze_panes = ws['d2']

# * FILL HEADER
ws['A1'] = 'День недели'
ws['A1'].alignment = Alignment(text_rotation=90)

ws['b1'] = 'Пара'
ws['b1'].alignment = Alignment(text_rotation=90)

ws['c1'] = 'Время\nзанятий'

for i, day in zip(range(2, 43, 8), WEEK_DAYS):
    ws.merge_cells(range_string=(f'A{i}:A{i+7}'))
    ws[f'a{i}'] = day
    ws[f'a{i}'].alignment = Alignment(
        horizontal='center', vertical='center', text_rotation=255)

for i in range(0, 48):
    ws[f'b{i+2}'] = i % 8 + 1
    ws[f'c{i+2}'] = LESSON_TIME[i % 8]

for i in range(0, 16):
    ws.cell(row=1, column=i+4).value = GROUPS[i]

# * CHANGE SIZES
ws.column_dimensions['A'].width = 3.3
ws.column_dimensions['B'].width = 3.3
ws.column_dimensions['C'].width = 12
for i in range(1, 50):
    ws.row_dimensions[i+1].height = 25

# * SET BORDERS
for i in range(1, 50, 8):
    set_border(ws, f'A{i}:S{i}', Border(bottom=Side(style='medium')))

for c in range(ord('C'), ord('S')+1):
    c = chr(c)
    set_border(ws, f'{c}1:{c}49', Border(right=Side(style='medium')))


# * HIDE UNNECESSARY COLUMNS
last_column = xl.utils.cell.column_index_from_string('XFD')
for idx in range(20, last_column+1):
    ws.column_dimensions[xl.utils.get_column_letter(idx)].hidden = True

# ws.row_dimensions.group(start=50, end=1048576, hidden=True)

wb.save('test.xlsx')
subprocess.Popen('open test.xlsx', shell=True)
